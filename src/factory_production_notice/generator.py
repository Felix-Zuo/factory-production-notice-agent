from __future__ import annotations

import html
import re
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .io_utils import write_json
from .models import ProductionNotice

CONTROL_CHAR_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")
FORMULA_PREFIXES = ("=", "+", "-", "@")


@dataclass
class GeneratedNotice:
    xlsx_path: Path
    html_path: Path
    manifest_path: Path
    agent_context_path: Path

    def as_manifest(self) -> dict[str, Any]:
        return {
            "output_dir": self.xlsx_path.parent.name or ".",
            "artifacts": {
                "xlsx": self.xlsx_path.name,
                "html": self.html_path.name,
                "manifest": self.manifest_path.name,
                "agent_context": self.agent_context_path.name,
            },
        }


def slug(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", value.strip()).strip("-")
    return cleaned or "notice"


def safe_excel_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    cleaned = CONTROL_CHAR_RE.sub("", value)
    if cleaned.lstrip().startswith(FORMULA_PREFIXES):
        return f"'{cleaned}"
    return cleaned


def write_cell(ws, row: int, column: int, value: Any):
    return ws.cell(row=row, column=column, value=safe_excel_value(value))


def generate_notice(payload: dict[str, Any], output_dir: str | Path) -> GeneratedNotice:
    notice = ProductionNotice.from_dict(payload)
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    basename = slug(f"{notice.notice_id}-{notice.product.item_code}")
    xlsx_path = out / f"{basename}.xlsx"
    html_path = out / f"{basename}.html"
    manifest_path = out / f"{basename}.manifest.json"
    agent_context_path = out / f"{basename}.agent_context.json"

    build_workbook(notice).save(xlsx_path)
    html_path.write_text(render_html(notice), encoding="utf-8")
    agent_context = notice.to_agent_context()
    write_json(agent_context_path, agent_context)
    write_json(
        manifest_path,
        {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "notice_id": notice.notice_id,
            "work_order": notice.work_order,
            "product_item_code": notice.product.item_code,
            "artifacts": {
                "xlsx": xlsx_path.name,
                "html": html_path.name,
                "agent_context": agent_context_path.name,
            },
        },
    )

    return GeneratedNotice(
        xlsx_path=xlsx_path,
        html_path=html_path,
        manifest_path=manifest_path,
        agent_context_path=agent_context_path,
    )


def build_workbook(notice: ProductionNotice) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Operations Notice"
    ws.sheet_view.showGridLines = False

    widths = [13, 18, 16, 14, 14, 14, 16, 16]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    title_fill = PatternFill("solid", fgColor="203040")
    section_fill = PatternFill("solid", fgColor="E8EEF5")
    label_fill = PatternFill("solid", fgColor="F6F8FB")
    thin = Side(style="thin", color="B8C0CC")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = notice.notice_type.upper()
    title.font = Font(color="FFFFFF", bold=True, size=18)
    title.fill = title_fill
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    rows = [
        ("Notice ID", notice.notice_id, "Work Order", notice.work_order, "Priority", notice.priority, "Due Date", notice.due_date),
        ("Requester", notice.customer, "Issuer", notice.issuer, "Quantity", f"{notice.quantity:g} {notice.quantity_unit}", "Revision", notice.product.revision),
        ("Subject ID", notice.product.item_code, "Subject", notice.product.name, "Category", notice.product.model, "", ""),
    ]
    start_row = 3
    for row_index, row_values in enumerate(rows, start=start_row):
        for col_index, value in enumerate(row_values, start=1):
            cell = write_cell(ws, row_index, col_index, value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_index % 2 == 1:
                cell.fill = label_fill
                cell.font = Font(bold=True)

    current = 7
    current = write_materials(ws, notice, current, section_fill, border)
    current = write_routing(ws, notice, current + 1, section_fill, border)
    current = write_packaging_quality(ws, notice, current + 1, section_fill, border)
    current = write_custom_fields(ws, notice, current + 1, section_fill, border)
    write_notes(ws, notice, current + 1, section_fill, border)

    for row in ws.iter_rows():
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "center"
            cell.alignment = alignment

    ws.freeze_panes = "A7"
    return wb


def write_section_header(ws, row: int, label: str, fill: PatternFill) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = write_cell(ws, row, 1, label)
    cell.fill = fill
    cell.font = Font(bold=True, color="203040")


def write_materials(ws, notice: ProductionNotice, start: int, fill: PatternFill, border: Border) -> int:
    write_section_header(ws, start, "Resource Requirement", fill)
    headers = ["Resource ID", "Resource", "Qty / Unit", "Required Qty", "Unit", "Source", "", ""]
    for col, value in enumerate(headers, start=1):
        cell = write_cell(ws, start + 1, col, value)
        cell.font = Font(bold=True)
        cell.border = border
    row = start + 2
    for line in notice.materials:
        values = [
            line.item_code,
            line.name,
            line.quantity_per,
            line.quantity_per * notice.quantity,
            line.unit,
            line.source,
            "",
            "",
        ]
        for col, value in enumerate(values, start=1):
            write_cell(ws, row, col, value).border = border
        row += 1
    return row


def write_routing(ws, notice: ProductionNotice, start: int, fill: PatternFill, border: Border) -> int:
    write_section_header(ws, start, "Execution Steps", fill)
    headers = ["Step", "Owner / Station", "Instruction", "Target Time Sec", "", "", "", ""]
    for col, value in enumerate(headers, start=1):
        cell = write_cell(ws, start + 1, col, value)
        cell.font = Font(bold=True)
        cell.border = border
    row = start + 2
    for step in notice.routing:
        values = [step.step, step.work_center, step.description, step.cycle_time_sec, "", "", "", ""]
        for col, value in enumerate(values, start=1):
            write_cell(ws, row, col, value).border = border
        row += 1
    return row


def write_packaging_quality(ws, notice: ProductionNotice, start: int, fill: PatternFill, border: Border) -> int:
    write_section_header(ws, start, "Fulfillment and Controls", fill)
    control_checks = notice.quality.get("critical_checks", notice.quality.get("critical_controls", []))
    rows = [
        ("Fulfillment Method", notice.packaging.get("method", "")),
        ("Units / Container", notice.packaging.get("units_per_carton", notice.packaging.get("units_per_container", ""))),
        ("Containers / Batch", notice.packaging.get("cartons_per_pallet", notice.packaging.get("containers_per_batch", ""))),
        ("Label Rule", notice.packaging.get("label_rule", "")),
        ("Gate Required", str(notice.quality.get("first_piece_required", notice.quality.get("approval_required", "")))),
        ("Sampling / Review Rule", notice.quality.get("sampling_rule", notice.quality.get("review_rule", ""))),
        ("Critical Checks", ", ".join(str(item) for item in control_checks)),
    ]
    row = start + 1
    for label, value in rows:
        write_cell(ws, row, 1, label).font = Font(bold=True)
        ws.cell(row=row, column=1).border = border
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        write_cell(ws, row, 2, value).border = border
        row += 1
    return row


def write_custom_fields(ws, notice: ProductionNotice, start: int, fill: PatternFill, border: Border) -> int:
    if not notice.custom_fields:
        return start - 1
    write_section_header(ws, start, "Template and Schedule Fields", fill)
    headers = ["Group", "Field", "Value", "", "", "", "", ""]
    for col, value in enumerate(headers, start=1):
        cell = write_cell(ws, start + 1, col, value)
        cell.font = Font(bold=True)
        cell.border = border
    row = start + 2
    for custom in notice.custom_fields:
        values = [custom.group, custom.label, custom.value, "", "", "", "", ""]
        for col, value in enumerate(values, start=1):
            write_cell(ws, row, col, value).border = border
        row += 1
    return row


def write_notes(ws, notice: ProductionNotice, start: int, fill: PatternFill, border: Border) -> int:
    write_section_header(ws, start, "Release Notes", fill)
    row = start + 1
    notes = notice.notes or ["Review and approve before releasing to operations."]
    for note in notes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        write_cell(ws, row, 1, f"- {note}").border = border
        row += 1
    return row


def render_html(notice: ProductionNotice) -> str:
    resource_rows = "\n".join(
        f"<tr><td>{esc(line.item_code)}</td><td>{esc(line.name)}</td><td>{line.quantity_per}</td>"
        f"<td>{line.quantity_per * notice.quantity:g}</td><td>{esc(line.unit)}</td><td>{esc(line.source)}</td></tr>"
        for line in notice.materials
    )
    step_rows = "\n".join(
        f"<tr><td>{esc(step.step)}</td><td>{esc(step.work_center)}</td><td>{esc(step.description)}</td>"
        f"<td>{step.cycle_time_sec:g}</td></tr>"
        for step in notice.routing
    )
    control_checks = notice.quality.get("critical_checks", notice.quality.get("critical_controls", []))
    checks = ", ".join(str(item) for item in control_checks)
    html_notes = notice.notes or ["Review and approve before releasing to operations."]
    notes = "\n".join(f"<li>{esc(note)}</li>" for note in html_notes)
    custom_section = render_custom_fields(notice)
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(notice.notice_id)} {esc(notice.notice_type)}</title>
  <style>
    :root {{
      --ink: #172033;
      --muted: #607086;
      --line: #ccd6e2;
      --blue: #183a5a;
      --paper: #ffffff;
      --soft: #eef3f8;
      --green: #2f7d5c;
      --amber: #a86613;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      padding: 32px;
      color: var(--ink);
      background: #edf2f6;
      font-family: Arial, "Microsoft YaHei", sans-serif;
    }}
    main {{
      max-width: 1180px;
      margin: 0 auto;
      background: var(--paper);
      border: 1px solid var(--line);
      box-shadow: 0 18px 48px rgba(23, 32, 51, .12);
    }}
    header {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 24px;
      align-items: center;
      padding: 24px 30px;
      color: white;
      background: linear-gradient(135deg, #14253a, #1d5274);
    }}
    h1 {{ margin: 0; font-size: 34px; letter-spacing: 0; }}
    .subtitle {{ margin-top: 8px; color: #cfe0ee; }}
    .status {{
      display: grid;
      gap: 8px;
      min-width: 220px;
      padding: 16px;
      border: 1px solid rgba(255,255,255,.28);
      background: rgba(255,255,255,.08);
    }}
    .status strong {{ font-size: 24px; color: #fff; }}
    .content {{ padding: 28px 30px 32px; }}
    .grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; }}
    .field {{ border: 1px solid var(--line); padding: 12px; background: #fbfdff; min-height: 74px; }}
    .label {{ color: var(--muted); font-size: 11px; font-weight: 700; letter-spacing: .06em; text-transform: uppercase; }}
    .value {{ margin-top: 8px; font-size: 16px; font-weight: 700; }}
    h2 {{
      margin: 28px 0 10px;
      padding: 10px 12px;
      color: #1b3149;
      background: var(--soft);
      border-left: 5px solid #346a92;
      font-size: 18px;
    }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ border: 1px solid var(--line); padding: 10px; text-align: left; vertical-align: top; }}
    th {{ background: #f5f8fb; color: #35445a; font-size: 12px; text-transform: uppercase; letter-spacing: .04em; }}
    .two {{ display: grid; grid-template-columns: 1fr 1fr; gap: 18px; }}
    .panel {{ border: 1px solid var(--line); padding: 14px; background: #fbfdff; min-height: 130px; }}
    .panel p {{ margin: 8px 0; color: var(--muted); line-height: 1.55; }}
    ul {{ margin: 8px 0 0; padding-left: 18px; color: var(--muted); line-height: 1.65; }}
    .approval {{
      display: grid;
      grid-template-columns: repeat(4, 1fr);
      gap: 10px;
      margin-top: 20px;
    }}
    .stamp {{
      min-height: 74px;
      border: 1px dashed #aab6c5;
      padding: 10px;
      color: var(--muted);
      background: #fafcff;
    }}
    @media (max-width: 860px) {{
      body {{ padding: 12px; }}
      header, .two, .grid, .approval {{ grid-template-columns: 1fr; }}
      .status {{ min-width: 0; }}
    }}
  </style>
</head>
<body>
  <main>
    <header>
      <div>
        <h1>{esc(notice.notice_type)}</h1>
        <div class="subtitle">Structured work package / operations coordination</div>
      </div>
      <div class="status">
        <span>Priority</span>
        <strong>{esc(notice.priority.upper())}</strong>
        <span>Due {esc(notice.due_date)}</span>
      </div>
    </header>
    <div class="content">
      <section class="grid">
        {field("Notice ID", notice.notice_id)}
        {field("Work Order", notice.work_order)}
        {field("Requester", notice.customer)}
        {field("Issuer", notice.issuer)}
        {field("Subject", notice.product.name)}
        {field("Subject ID", notice.product.item_code)}
        {field("Category / Revision", f"{notice.product.model} / {notice.product.revision}")}
        {field("Quantity", f"{notice.quantity:g} {notice.quantity_unit}")}
      </section>
      <h2>Resource Requirement</h2>
      <table><thead><tr><th>Resource ID</th><th>Resource</th><th>Qty / Unit</th><th>Required Qty</th><th>Unit</th><th>Source</th></tr></thead><tbody>{resource_rows}</tbody></table>
      <h2>Execution Steps</h2>
      <table><thead><tr><th>Step</th><th>Owner / Station</th><th>Instruction</th><th>Target Time Sec</th></tr></thead><tbody>{step_rows}</tbody></table>
      <section class="two">
        <div>
          <h2>Fulfillment</h2>
          <div class="panel">
            <p><strong>Method:</strong> {esc(str(notice.packaging.get("method", "")))}</p>
            <p><strong>Container:</strong> {esc(str(notice.packaging.get("units_per_carton", notice.packaging.get("units_per_container", ""))))} units / container</p>
            <p><strong>Batch:</strong> {esc(str(notice.packaging.get("cartons_per_pallet", notice.packaging.get("containers_per_batch", ""))))} containers / batch</p>
            <p><strong>Label:</strong> {esc(str(notice.packaging.get("label_rule", "")))}</p>
          </div>
        </div>
        <div>
          <h2>Controls</h2>
          <div class="panel">
            <p><strong>Gate:</strong> {esc(str(notice.quality.get("first_piece_required", notice.quality.get("approval_required", ""))))}</p>
            <p><strong>Review:</strong> {esc(str(notice.quality.get("sampling_rule", notice.quality.get("review_rule", ""))))}</p>
            <p><strong>Checks:</strong> {esc(checks)}</p>
          </div>
        </div>
      </section>
      {custom_section}
      <h2>Release Notes</h2>
      <ul>{notes}</ul>
      <section class="approval">
        <div class="stamp">Planning</div>
        <div class="stamp">Owner</div>
        <div class="stamp">Controls</div>
        <div class="stamp">Fulfillment</div>
      </section>
    </div>
  </main>
</body>
</html>
"""


def render_custom_fields(notice: ProductionNotice) -> str:
    if not notice.custom_fields:
        return ""
    rows = "\n".join(
        f"<tr><td>{esc(custom.group)}</td><td>{esc(custom.label)}</td><td>{esc(custom.value)}</td></tr>"
        for custom in notice.custom_fields
    )
    return f"""
      <h2>Template and Schedule Fields</h2>
      <table><thead><tr><th>Group</th><th>Field</th><th>Value</th></tr></thead><tbody>{rows}</tbody></table>
"""


def field(label: str, value: str) -> str:
    return f'<div class="field"><div class="label">{esc(label)}</div><div class="value">{esc(str(value))}</div></div>'


def esc(value: str) -> str:
    return html.escape(value or "")
