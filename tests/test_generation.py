from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from factory_production_notice.agent_contract import build_agent_interface
from factory_production_notice.csv_adapter import CsvImportError, import_csv_notices
from factory_production_notice.generator import generate_notice
from factory_production_notice.io_utils import read_json
from factory_production_notice.models import NoticeValidationError, ProductionNotice
from factory_production_notice.profiles import profile_catalog
from factory_production_notice.server import is_loopback_host
from factory_production_notice.validation import validate_notice_payload
from scripts.package_project import should_skip


def test_demo_generation(tmp_path: Path) -> None:
    payload = read_json("sample_data/demo_notice_request.json")
    result = generate_notice(payload, tmp_path)

    assert result.xlsx_path.exists()
    assert result.html_path.exists()
    assert result.manifest_path.exists()
    assert result.agent_context_path.exists()
    manifest = result.as_manifest()
    assert manifest["output_dir"] == tmp_path.name
    assert manifest["artifacts"]["html"] == result.html_path.name
    assert str(tmp_path) not in manifest["artifacts"]["html"]

    context = read_json(result.agent_context_path)
    assert context["notice_id"] == "ON-2026-DEMO-001"
    assert context["domain"] == "warehouse-fulfillment"
    assert context["required_resources"][0]["required_quantity"] == 240
    assert context["required_materials"][0]["required_quantity"] == 240


def test_agent_interface_shape() -> None:
    spec = build_agent_interface()
    capability_names = {item["name"] for item in spec["capabilities"]}
    assert "generate_operations_notice" in capability_names
    assert "generate_notice" in capability_names
    assert "validate_notice" in capability_names
    assert "import_csv_work_packages" in capability_names
    assert "list_scenario_profiles" in capability_names
    assert spec["input_contract"]["sample_path"].endswith("demo_notice_request.json")


def test_profile_catalog_lists_cross_domain_profiles() -> None:
    catalog = profile_catalog()
    profile_ids = {item["id"] for item in catalog["profiles"]}

    assert catalog["count"] == 5
    assert "warehouse-fulfillment" in profile_ids
    assert "maintenance-service" in profile_ids
    assert "compliance-review" in profile_ids


def test_csv_adapter_imports_batch_sample() -> None:
    notices = import_csv_notices("sample_data/csv/work_package_notices.csv")

    assert len(notices) == 2
    assert notices[0]["domain"] == "maintenance-service"
    assert notices[0]["subject"]["subject_id"] == "ASSET-PUMP-LINE-2"
    assert len(notices[0]["resources"]) == 3
    assert len(notices[0]["steps"]) == 3
    assert notices[0]["controls"]["approval_required"] is True
    assert "lockout" in notices[0]["controls"]["critical_controls"]


def test_csv_adapter_rejects_invalid_boolean(tmp_path: Path) -> None:
    source = tmp_path / "bad.csv"
    source.write_text(
        "\n".join(
            [
                "notice_id,work_order,quantity,due_date,subject_id,subject_name,approval_required",
                "ON-BAD-CSV,WO-BAD,1,2026-06-30,SUBJ,Bad CSV,maybe",
            ]
        ),
        encoding="utf-8",
    )

    try:
        import_csv_notices(source)
    except CsvImportError as exc:
        assert "approval_required must be true or false" in str(exc)
    else:
        raise AssertionError("Expected CsvImportError")


def test_legacy_manufacturing_payload_still_works(tmp_path: Path) -> None:
    payload = {
        "notice_id": "PN-LEGACY-001",
        "work_order": "WO-LEGACY-001",
        "quantity": 12,
        "due_date": "2026-06-30",
        "product": {"item_code": "FG-001", "name": "Legacy Assembly"},
        "materials": [{"item_code": "RM-001", "name": "Legacy Part", "quantity_per": 2}],
        "routing": [{"step": "OP10", "work_center": "Cell A", "description": "Build"}],
    }

    result = generate_notice(payload, tmp_path)
    context = read_json(result.agent_context_path)

    assert context["product"]["item_code"] == "FG-001"
    assert context["required_materials"][0]["required_quantity"] == 24


def test_rejects_invalid_quantity() -> None:
    payload = {
        "notice_id": "ON-BAD-001",
        "work_order": "OPS-BAD-001",
        "quantity": 0,
        "due_date": "2026-06-30",
        "subject": {"subject_id": "BAD", "name": "Invalid Quantity"},
    }

    try:
        ProductionNotice.from_dict(payload)
    except NoticeValidationError as exc:
        assert "quantity must be greater than zero" in str(exc)
    else:
        raise AssertionError("Expected NoticeValidationError")


def test_rejects_non_object_resource_entries() -> None:
    payload = {
        "notice_id": "ON-BAD-002",
        "work_order": "OPS-BAD-002",
        "quantity": 1,
        "due_date": "2026-06-30",
        "subject": {"subject_id": "BAD", "name": "Invalid Resource"},
        "resources": ["not an object"],
    }

    try:
        ProductionNotice.from_dict(payload)
    except NoticeValidationError as exc:
        assert "resources/materials entries must be objects" in str(exc)
    else:
        raise AssertionError("Expected NoticeValidationError")


def test_validate_notice_payload_catches_bad_nested_numbers() -> None:
    result = validate_notice_payload(
        {
            "notice_id": "ON-BAD-003",
            "work_order": "OPS-BAD-003",
            "quantity": 1,
            "due_date": "2026-06-30",
            "subject": {"subject_id": "BAD", "name": "Bad Nested Number"},
            "resources": [{"resource_id": "R1", "name": "Resource", "quantity_per": "many"}],
        }
    )

    assert not result.ok
    assert result.errors == ["quantity_per must be a number"]


def test_package_excludes_generated_probe_outputs() -> None:
    assert should_skip(Path("output") / "demo.html")
    assert should_skip(Path("output_http_probe") / "demo.html")
    assert should_skip(Path("output_alias_probe") / "demo.html")


def test_excel_output_escapes_formula_like_user_text(tmp_path: Path) -> None:
    payload = {
        "notice_id": "ON-SAFE-001",
        "work_order": "OPS-SAFE-001",
        "quantity": 1,
        "due_date": "2026-06-30",
        "requester": "=cmd|' /C calc'!A0",
        "subject": {"subject_id": "SAFE-001", "name": "+malicious"},
        "resources": [{"resource_id": "RES-001", "name": "@payload", "quantity_per": 1}],
        "steps": [{"step": "S10", "owner": "-owner", "instruction": "=formula"}],
    }

    result = generate_notice(payload, tmp_path)
    wb = load_workbook(result.xlsx_path, data_only=False)
    ws = wb.active
    values = [cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str)]

    assert "'=cmd|' /C calc'!A0" in values
    assert "'+malicious" in values
    assert "'@payload" in values
    assert "'-owner" in values
    assert "'=formula" in values


def test_validate_notice_payload_reports_summary_and_warnings() -> None:
    result = validate_notice_payload(
        {
            "notice_id": "ON-WARN-001",
            "work_order": "OPS-WARN-001",
            "quantity": 1,
            "due_date": "2026-06-30",
            "subject": {"subject_id": "WARN-001", "name": "Sparse payload"},
        }
    )

    assert result.ok
    assert result.summary["notice_id"] == "ON-WARN-001"
    assert "No resources/materials are listed." in result.warnings


def test_loopback_host_guard() -> None:
    assert is_loopback_host("127.0.0.1")
    assert is_loopback_host("localhost")
    assert is_loopback_host("::1")
    assert not is_loopback_host("0.0.0.0")
    assert not is_loopback_host("192.168.1.10")
